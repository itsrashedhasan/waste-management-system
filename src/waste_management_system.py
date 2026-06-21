from openpyxl import Workbook, load_workbook


class Waste:
    def __init__(self, waste_id, waste_type, descriptions):
        self.waste_id = waste_id
        self.waste_type = waste_type
        self.descriptions = descriptions

    def display_waste_details(self):
        return f"ID: {self.waste_id}, Type: {self.waste_type}, Descriptions: {self.descriptions}"


class CollectionSchedule:
    def __init__(self, schedule_id, waste, collection_date, location):
        self.schedule_id = schedule_id
        self.waste = waste
        self.collection_date = collection_date
        self.location = location

    def display_schedule_details(self):
        return f"ID: {self.schedule_id}, Waste Type: {self.waste.waste_type}, Collection Date: {self.collection_date}, Location: {self.location}"


class WasteCollector:
    def __init__(self, collector_id, name):
        self.collector_id = collector_id
        self.name = name
        self.assigned_routes = []

    def assign_route(self, schedule):
        self.assigned_routes.append(schedule)

    def display_collector_details(self):
        details = f"Collector ID: {self.collector_id}, Name: {self.name}, Routes:\n"
        for route in self.assigned_routes:
            details += route.display_schedule_details() + "\n"
        return details


class WasteManagementSystem:
    def __init__(self):
        self.wastes = []
        self.schedules = []
        self.collectors = []
        self.load()

    def load(self):
        # Loading wastes from the "wastes.xlsx" file
        try:
            self.workbook = load_workbook("wastes.xlsx")  # Loading the workbook
            self.sheet = self.workbook.active  # Getting the active sheet of the workbook
            for row in self.sheet.iter_rows(min_row=2, values_only=True):  # Iterating over rows (min_row=2 skips header)
                self.wastes.append(Waste(row[0], row[1], row[2]))  # Creating a Waste object and adding it to the list
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook and save it with a header
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Waste Sheet"
            self.sheet.append(["Waste ID", "Type", "Description"])  # Adding headers
            self.workbook.save("wastes.xlsx")  # Saving the workbook

        # Loading schedules from "schedules.xlsx" file
        try:
            self.workbook = load_workbook("schedules.xlsx")
            self.sheet = self.workbook.active
            for row in self.sheet.iter_rows(min_row=2, values_only=True):  # Iterating over rows
                for waste in self.wastes:  # Finding the corresponding waste for each schedule
                    if waste.waste_id == row[1]:
                        self.schedules.append(
                            CollectionSchedule(row[0], waste, row[2], row[3])  # Creating a schedule object and adding it to the list
                        )
                        break
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook and save it with a header
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Schedules Sheet"
            self.sheet.append(["Schedule ID", "Waste ID", "Collection Date", "Location"])
            self.workbook.save("schedules.xlsx")

        # Loading collectors from "collectors.xlsx" file
        try:
            self.workbook = load_workbook("collectors.xlsx")
            self.sheet = self.workbook.active
            for row in self.sheet.iter_rows(min_row=2, values_only=True):  # Iterating over rows
                self.collectors.append(WasteCollector(row[0], row[1]))  # Creating a WasteCollector object and adding it to the list
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook and save it with a header
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Collector Sheet"
            self.sheet.append(["Collector ID", "Name"])
            self.workbook.save("collectors.xlsx")

    def save(self):
        # Saving wastes to "wastes.xlsx"
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.append(["Waste ID", "Type", "Description"])  # Adding headers
        for waste in self.wastes:  # Saving each waste item in the list to the sheet
            self.sheet.append([waste.waste_id, waste.waste_type, waste.descriptions])
        self.workbook.save("wastes.xlsx")  # Saving the workbook

        # Saving schedules to "schedules.xlsx"
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.append(["Schedule ID", "Waste ID", "Collection Date", "Location"])  # Adding headers
        for schedule in self.schedules:  # Saving each schedule item in the list to the sheet
            self.sheet.append(
                [
                    schedule.schedule_id,
                    schedule.waste.waste_id,
                    schedule.collection_date,
                    schedule.location,
                ]
            )
        self.workbook.save("schedules.xlsx")

        # Saving collectors to "collectors.xlsx"
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.append(["Collector ID", "Name"])  # Adding headers
        for collector in self.collectors:  # Saving each collector item in the list to the sheet
            self.sheet.append([collector.collector_id, collector.name])
        self.workbook.save("collectors.xlsx")

    def add_waste(self):
        # Prompting user to add a waste
        waste_id = input("Enter Waste ID: ")
        for waste in self.wastes:  # Checking if the waste ID already exists
            if waste.waste_id == waste_id:
                print(f"The waste ID {waste_id} already exists. Please use another ID.")
                return
        waste_type = input("Enter Waste Type (Organic/Recyclable/Hazardous): ")
        description = input("Enter Description: ")
        self.wastes.append(Waste(waste_id, waste_type, description))  # Adding the waste to the list
        print("Waste type added successfully!")
        self.save()
    def schedule_collection(self):
        # Collects details to schedule a new waste collection
        schedule_id = input("Enter Schedule ID: ")
        # Checks if the schedule ID already exists
        for schedule in self.schedules:
            if schedule.schedule_id == schedule_id:
                print(f"The Schedule ID {schedule_id} already exists. Please use another ID.")
                return
        # Inputs for waste ID, collection date, and location
        waste_id = input("Enter Waste ID: ")
        collection_date = input("Enter Collection Date (YYYY-MM-DD): ")
        location = input("Enter Location: ")
        
        selected_waste = None
        # Finds the waste with the given ID
        for waste_item in self.wastes:
            if waste_item.waste_id == waste_id:
                selected_waste = waste_item
                break
        if selected_waste:
            # Adds a new collection schedule if waste is valid
            self.schedules.append(
                CollectionSchedule(schedule_id, selected_waste, collection_date, location)
            )
            print("Collection Scheduled Successfully!")
        else:
            # Handles invalid waste ID input
            print("Invalid Waste ID! Please check the entered ID and try again.")

        self.save()  # Saves the changes for persistence

    def assign_collector(self):
        # Assigns a collector to a specific schedule
        collector_id = input("Enter Collector ID: ")
        collector_name = input("Enter Collector Name: ")
        schedule_id = input("Enter Schedule ID to assign: ")
        
        selected_schedule = None
        # Finds the schedule by ID
        for schedule_item in self.schedules:
            if schedule_item.schedule_id == schedule_id:
                selected_schedule = schedule_item
                break
        if not selected_schedule:
            # Handles invalid schedule ID input
            print("Invalid Schedule ID! Please check and try again.")
            return
        
        selected_collector = None
        # Finds the collector by ID
        for collector_item in self.collectors:
            if collector_item.collector_id == collector_id:
                selected_collector = collector_item
                break
        if not selected_collector:
            # Creates a new collector if it does not exist
            selected_collector = WasteCollector(collector_id, collector_name)
            self.collectors.append(selected_collector)
        # Assigns the collector to the selected schedule
        selected_collector.assign_route(selected_schedule)
        print("Collector Assigned Successfully!")

        self.save()  # Saves the changes for persistence

    def display_schedules(self):
        # Displays all collection schedules
        print("Collection Schedules:")
        for schedule in self.schedules:
            # Uses the schedule's display method for details
            print(schedule.display_schedule_details())

    def delete_waste(self):
        # Deletes a waste by its ID
        waste_id = input("Enter the Waste ID to delete: ")
        waste_to_delete = None
        # Searches for the waste by ID
        for waste in self.wastes:
            if waste.waste_id == waste_id:
                waste_to_delete = waste
                break
        
        if waste_to_delete:
            # Removes the waste if found
            self.wastes.remove(waste_to_delete)
            print(f"Waste with ID {waste_id} deleted successfully!")
        else:
            # Handles invalid waste ID input
            print("Waste ID not found!")

    def delete_schedule(self):
        # Deletes a schedule by its ID
        schedule_id = input("Enter the Schedule ID to delete: ")
        schedule_to_delete = None
        # Searches for the schedule by ID
        for schedule in self.schedules:
            if schedule.schedule_id == schedule_id:
                schedule_to_delete = schedule
                break
        
        if schedule_to_delete:
            # Removes the schedule if found
            self.schedules.remove(schedule_to_delete)
            print(f"Schedule with ID {schedule_id} deleted successfully!")
        else:
            # Handles invalid schedule ID input
            print("Schedule ID not found!")

    def delete_collector(self):
        # Deletes a collector by their ID
        collector_id = input("Enter the Collector ID to delete: ")
        collector_to_delete = None
        # Searches for the collector by ID
        for collector in self.collectors:
            if collector.collector_id == collector_id:
                collector_to_delete = collector
                break
        
        if collector_to_delete:
            # Removes the collector if found
            self.collectors.remove(collector_to_delete)
            print(f"Collector with ID {collector_id} deleted successfully!")
        else:
            # Handles invalid collector ID input
            print("Collector ID not found!")

    def update_excel(self):
        # Exports wastes data to an Excel file
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.append(["Waste ID", "Type", "Description"])
        for waste in self.wastes:
            self.sheet.append([waste.waste_id, waste.waste_type, waste.description])
        self.workbook.save("wastes.xlsx")

        # Exports schedules data to an Excel file
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.append(["Schedule ID", "Waste ID", "Collection Date", "Location"])
        for schedule in self.schedules:
            self.sheet.append([schedule.schedule_id, schedule.waste.waste_id, schedule.collection_date, schedule.location])
        self.workbook.save("schedules.xlsx")

        # Exports collectors data to an Excel file
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.append(["Collector ID", "Name"])
        for collector in self.collectors:
            self.sheet.append([collector.collector_id, collector.name])
        self.workbook.save("collectors.xlsx")

    def menu(self):
        # Main menu loop
        while True:
            print("\nWelcome to the Waste Management System!")
            print("1. Add Waste Type")
            print("2. Schedule Collection")
            print("3. Assign Collector")
            print("4. View Collection Schedules")
            print("5. Delete Waste")
            print("6. Delete Schedule")
            print("7. Delete Collector")
            print("8. Exit")
            choice = input("Choose an option: ")

            # Routes user input to the appropriate method
            if choice == "1":
                self.add_waste()
            elif choice == "2":
                self.schedule_collection()
            elif choice == "3":
                self.assign_collector()
            elif choice == "4":
                self.display_schedules()
            elif choice == "5":
                self.delete_waste()
            elif choice == "6":
                self.delete_schedule()
            elif choice == "7":
                self.delete_collector()
            elif choice == "8":
                self.save()  # Saves data before exiting
                print("Exiting... Goodbye")
                break
            else:
                print("Invalid choice! Please try again.")

if __name__ == "__main__":
    # Initializes the system and starts the menu
    obj_system = WasteManagementSystem()
    obj_system.menu()
